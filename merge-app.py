import streamlit as st
import pandas as pd
import os
import tempfile
import shutil
from datetime import datetime
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import copy

# Konfigurasi halaman
st.set_page_config(
    page_title="Excel File Merger",
    page_icon="📊",
    layout="wide"
)

MAX_MISSING_VALUES = 10
required_columns = [
    'PERIODE', 'KANTOR WILAYAH', 'KANTOR CABANG', 'UNIT KERJA', 'LOAN TYPE',
    'CIFNO', 'NO REKENING', 'NAMA DEBITUR', 'STATUS REKENING', 'STATUS DATE',
    'NILAI TERCATAT', 'CKPN SEBELUM', 'CKPN BERJALAN', 'BIAYA CKPN',
    'FLAG RESTRUK', 'STAGE', 'KOLEKTIBILITAS', 'UMUR TUNGGAKAN', 'SEGMENTASI'
]
numeric_columns = ['NILAI TERCATAT', 'CKPN SEBELUM', 'CKPN BERJALAN', 'BIAYA CKPN']

def format_date_to_indonesian(date_obj):
    """Convert date object to format: dd Nam yyyy"""
    month_names = {
        1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr',
        5: 'May', 6: 'Jun', 7: 'Jul', 8: 'Aug',
        9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'
    }
    
    day = date_obj.day
    month = month_names[date_obj.month]
    year = date_obj.year
    
    return f"{day} {month} {year}"

def format_numeric_value(value):
    if pd.isna(value) or value == '' or value == 'nan':
        return value
    
    try:
        if isinstance(value, str):
            clean_value = value.replace(',', '')
            numeric_value = float(clean_value)
        else:
            numeric_value = float(value)
        
        if numeric_value == int(numeric_value):
            return f"{int(numeric_value):,}"
        else:
            return f"{numeric_value:,.2f}"
            
    except (ValueError, TypeError):
        return value

def is_missing_value(value):
    if pd.isna(value):
        return True
    if isinstance(value, str):
        stripped = value.strip()
        return stripped == '' or stripped.lower() == 'nan'
    return False

def count_missing_values(row):
    return sum(1 for value in row if is_missing_value(value))

def copy_cell_style(source_cell, target_cell):
    """Copy style from source cell to target cell"""
    try:
        if source_cell.fill and source_cell.fill.fill_type:
            target_cell.fill = copy.copy(source_cell.fill)
        if source_cell.font:
            target_cell.font = copy.copy(source_cell.font)
        if source_cell.border:
            target_cell.border = copy.copy(source_cell.border)
        if source_cell.alignment:
            target_cell.alignment = copy.copy(source_cell.alignment)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
    except Exception as e:
        pass  # Ignore styling errors

def copy_row_styles(source_ws, target_ws, source_row, target_row, max_col):
    """Copy entire row styles from source to target"""
    try:
        for col in range(1, max_col + 1):
            source_cell = source_ws.cell(row=source_row, column=col)
            target_cell = target_ws.cell(row=target_row, column=col)
            copy_cell_style(source_cell, target_cell)
    except Exception as e:
        pass  # Ignore styling errors

def preserve_template_formatting(template_path, output_path, data_start_row=15):
    """Preserve all formatting from template including backgrounds"""
    try:
        # Load template to get original formatting
        template_wb = load_workbook(template_path)
        template_ws = template_wb.active
        
        # Load output file
        output_wb = load_workbook(output_path)
        output_ws = output_wb.active
        
        # Copy all formatting from template (including backgrounds)
        max_row = template_ws.max_row
        max_col = template_ws.max_column
        
        # Copy all cell styles from template
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                template_cell = template_ws.cell(row=row, column=col)
                output_cell = output_ws.cell(row=row, column=col)
                
                # Copy value if it's not in data area
                if row < data_start_row:
                    output_cell.value = template_cell.value
                
                # Copy style
                copy_cell_style(template_cell, output_cell)
        
        # Copy column widths
        for col in template_ws.column_dimensions:
            if col in output_ws.column_dimensions:
                output_ws.column_dimensions[col].width = template_ws.column_dimensions[col].width
        
        # Copy row heights
        for row in template_ws.row_dimensions:
            if row in output_ws.row_dimensions:
                output_ws.row_dimensions[row].height = template_ws.row_dimensions[row].height
        
        # Save the formatted output
        output_wb.save(output_path)
        
        template_wb.close()
        output_wb.close()
        
        return True, "Formatting berhasil dipertahankan"
        
    except Exception as e:
        return False, f"Error preserving formatting: {e}"

def update_template_metadata(template_path, metadata):
    try:
        # Try xlwings first
        app = xw.App(visible=False)
        wb = app.books.open(template_path)
        ws = wb.sheets[0]
        
        ws.range('F5').value = metadata['periode']
        ws.range('F7').value = metadata['kanwil']
        ws.range('F9').value = metadata['kanca']
        ws.range('F11').value = metadata['unit_kerja']
        
        wb.save()
        wb.close()
        app.quit()
        return True, "Metadata berhasil diinput"
        
    except Exception as e:
        try:
            # Fallback to openpyxl
            wb = load_workbook(template_path)
            ws = wb.active
            
            ws['F5'] = metadata['periode']
            ws['F7'] = metadata['kanwil']
            ws['F9'] = metadata['kanca']
            ws['F11'] = metadata['unit_kerja']
            
            wb.save(template_path)
            wb.close()
            return True, "Metadata berhasil diinput dengan fallback"
            
        except Exception as e2:
            return False, f"Error updating metadata: {e2}"

def process_files(uploaded_files, temp_dir):
    data_list = []
    processing_log = []
    
    for uploaded_file in uploaded_files:
        file_path = os.path.join(temp_dir, uploaded_file.name)
        with open(file_path, 'wb') as f:
            f.write(uploaded_file.getbuffer())
        
        try:
            df = pd.read_excel(file_path, sheet_name='BIAYA_CKPN_UKER', header=13, dtype=str)
            df.columns = df.columns.str.strip()
            
            missing_cols = [col for col in required_columns if col not in df.columns]
            if missing_cols:
                processing_log.append(f"⚠️ File {uploaded_file.name}: Kolom yang hilang: {missing_cols}")
                continue

            initial_rows = len(df)
            
            df_required = df[required_columns]
            missing_counts = df_required.apply(count_missing_values, axis=1)
            
            valid_rows_mask = missing_counts <= MAX_MISSING_VALUES
            df_filtered = df[valid_rows_mask].copy()
            
            processing_log.append(f"File {uploaded_file.name}: {initial_rows} baris awal -> {len(df_filtered)} baris")

            for col in numeric_columns:
                if col in df_filtered.columns:
                    df_filtered.loc[:, col] = df_filtered[col].apply(format_numeric_value)
            
            if not df_filtered.empty:
                data_list.append(df_filtered)
                
        except Exception as e:
            processing_log.append(f"Gagal membaca {uploaded_file.name}: {e}")
    
    return data_list, processing_log

def create_final_file(data_list, template_path, output_path):
    if not data_list:
        return False, "Tidak ada data untuk digabung!"

    gabungan_df = pd.concat(data_list, ignore_index=True)

    initial_combined_rows = len(gabungan_df)
    gabungan_df_required = gabungan_df[required_columns]
    missing_counts_combined = gabungan_df_required.apply(count_missing_values, axis=1)
    valid_combined_mask = missing_counts_combined <= MAX_MISSING_VALUES
    gabungan_df = gabungan_df[valid_combined_mask].copy()

    for col in numeric_columns:
        if col in gabungan_df.columns:
            gabungan_df.loc[:, col] = gabungan_df[col].apply(format_numeric_value)

    # Copy template to output
    shutil.copy2(template_path, output_path)

    data_start_row = 15
    
    try:
        # Try xlwings first
        app = xw.App(visible=False)
        wb = app.books.open(output_path)
        ws = wb.sheets[0]
        
        for r_idx, row in enumerate(gabungan_df.itertuples(index=False), start=data_start_row):
            for c_idx, value in enumerate(row, start=1):
                if is_missing_value(value):
                    continue
                ws.range(r_idx, c_idx).value = value
        
        wb.save()
        wb.close()
        app.quit()
        
        return True, f"✅ Data berhasil digabung ke dalam file. Total {len(gabungan_df)} baris data"
        
    except Exception as e:
        try:
            # Fallback to openpyxl with enhanced formatting preservation
            wb = load_workbook(output_path)
            ws = wb.active
            
            # Get template formatting reference
            template_wb = load_workbook(template_path)
            template_ws = template_wb.active
            
            # Insert data while preserving formatting
            for r_idx, row in enumerate(gabungan_df.itertuples(index=False), start=data_start_row):
                for c_idx, value in enumerate(row, start=1):
                    if is_missing_value(value):
                        continue
                    
                    # Set the value
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.value = value
                    
                    # Try to copy style from template row if it exists
                    if r_idx <= template_ws.max_row:
                        template_cell = template_ws.cell(row=r_idx, column=c_idx)
                        copy_cell_style(template_cell, cell)
                    elif data_start_row <= template_ws.max_row:
                        # Use style from header row or first data row
                        template_cell = template_ws.cell(row=data_start_row, column=c_idx)
                        copy_cell_style(template_cell, cell)
            
            wb.save(output_path)
            wb.close()
            template_wb.close()
            
            # Additional step: Preserve all template formatting
            preserve_success, preserve_msg = preserve_template_formatting(template_path, output_path, data_start_row)
            
            return True, f"✅ Data berhasil digabung (fallback mode). Total {len(gabungan_df)} baris data"
            
        except Exception as e2:
            return False, f"Error: {e2}"

def main():
    st.title("📊 Excel File Merger - BRI CKPN")
    st.markdown("---")
    
    # Upload Template
    st.header("1. 📁 Upload Template File")
    template_file = st.file_uploader(
        "Pilih file template Excel (.xlsx)",
        type=['xlsx'],
        help="Upload file template yang akan digunakan sebagai dasar formatting"
    )
    
    st.header("2. 📝 Input Metadata")
    col1, col2 = st.columns(2)
    
    with col1:
        periode_date = st.date_input(
            "Periode",
            value=datetime.now(),
            help="Pilih tanggal periode (akan diformat sebagai dd Nam yyyy)"
        )
        
        kanwil = st.text_input("Kantor Wilayah", value="", help="Contoh: G - Semarang")
    
    with col2:
        kanca = st.text_input("Kantor Cabang", value="", help="Contoh: 00156 - KC Batang")
        unit_kerja = st.text_input("Unit Kerja", value="", help="Contoh: 00156 - KC Batang")
    
    periode_formatted = format_date_to_indonesian(periode_date) if periode_date else ""
    
    metadata = {
        'periode': periode_formatted,
        'kanwil': kanwil,
        'kanca': kanca,
        'unit_kerja': unit_kerja
    }
    
    st.header("3. 📤 Upload Files untuk Digabung")
    uploaded_files = st.file_uploader(
        "Pilih file Excel yang akan digabungkan",
        type=['xlsx'],
        accept_multiple_files=True,
        help="Upload satu atau lebih file Excel yang akan digabungkan"
    )
    
    st.header("4. 🔄 Proses dan Download")
    
    if st.button("🚀 Proses File", type="primary"):
        if template_file is None:
            st.error("❌ Silakan upload file template terlebih dahulu!")
            return
        
        if not uploaded_files:
            st.error("❌ Silakan upload minimal satu file untuk digabungkan!")
            return
        
        if not all([periode_date, kanwil, kanca, unit_kerja]):
            st.error("❌ Silakan lengkapi semua metadata!")
            return
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        with tempfile.TemporaryDirectory() as temp_dir:
            template_path = os.path.join(temp_dir, 'template.xlsx')
            with open(template_path, 'wb') as f:
                f.write(template_file.getbuffer())
            
            progress_bar.progress(20)
            status_text.text("🔄 Menginput metadata...")
            
            success, msg = update_template_metadata(template_path, metadata)
            if not success:
                st.error(msg)
                return
            
            progress_bar.progress(40)
            status_text.text("🔄 Memproses file Excel...")
            
            data_list, processing_log = process_files(uploaded_files, temp_dir)

            progress_bar.progress(70)
            status_text.text("🔄 Menggabungkan data...")

            output_path = os.path.join(temp_dir, 'hasil_gabungan.xlsx')
            success, result_message = create_final_file(data_list, template_path, output_path)

            if success:
                progress_bar.progress(90)
                status_text.text("🔄 Mempertahankan formatting...")
                
                # Final formatting preservation step
                preserve_success, preserve_msg = preserve_template_formatting(template_path, output_path)
                
                progress_bar.progress(100)
                status_text.text("✅ Pemrosesan selesai!")
                
                with open(output_path, 'rb') as f:
                    file_data = f.read()

                filename = f"hasil_gabungan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                st.download_button(
                    label="💾 Download File Hasil",
                    data=file_data,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )
                
                # Show processing log
                if processing_log:
                    with st.expander("📋 Log Pemrosesan"):
                        for log in processing_log:
                            st.write(log)
                            
            else:
                st.error(result_message)

if __name__ == "__main__":
    main()